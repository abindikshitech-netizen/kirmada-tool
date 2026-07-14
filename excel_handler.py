import pandas as pd
import os
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from models.shop import Shop
from logger import app_logger

class ExcelHandler:
    def __init__(self, input_path: str, output_path: str):
        self.input_path = input_path
        self.output_path = output_path
        self.df = None

    def read_input(self) -> Tuple[List[Shop], int]:
        try:
            self.df = pd.read_excel(self.input_path)
            
            col_name = "Shop Name + Old Address" if "Shop Name + Old Address" in self.df.columns else self.df.columns[0]
            col_dist = "District" if "District" in self.df.columns else self.df.columns[1]

            shops = []
            for index, row in self.df.iterrows():
                shop = Shop(
                    original_shop_name_address=str(row[col_name]).strip() if pd.notna(row[col_name]) else "",
                    original_district=str(row[col_dist]).strip() if pd.notna(row[col_dist]) else "",
                    row_index=index
                )
                shops.append(shop)
            
            app_logger.info(f"Loaded {len(shops)} shops from {self.input_path}")
            return shops, len(shops)
        except Exception as e:
            app_logger.error(f"Failed to read input Excel: {e}")
            raise

    def _format_sheet(self, ws):
        if ws.max_row <= 1: return
        for cell in ws[1]: cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    def write_output(self, shops: List[Shop]):
        app_logger.info(f"Audit: Received for Excel Export: {len(shops)}")
        if not shops: return
        try:
            columns_order = [
                "Shop Name + Old Address", "District", "Pincode", "Latest Complete Address",
                "Phone Number", "Business Status", "Verification Status", "Confidence Score",
                "Data Source", "Latitude", "Longitude", "Website",
                "Business Name", "Business Category", "Open/Closed Status", "Ratings", "Review Count", "Google Maps URL", "Plus Code",
                "Place ID", "Error Message"
            ]

            wb = Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
                
            all_data = [shop.to_dict() for shop in shops]
            for i, shop in enumerate(shops): all_data[i]["Error Message"] = shop.error_message
                
            df_all = pd.DataFrame(all_data)
            for col in columns_order:
                if col not in df_all.columns: df_all[col] = ""
            df_all = df_all[columns_order]
            
            df_failed = df_all[df_all["Error Message"] != ""]
            df_manual = df_all[(df_all["Verification Status"] == "Manual Review") & (df_all["Error Message"] == "")]
            df_verified = df_all[(df_all["Verification Status"].isin(["Verified", "Likely Match"])) & (df_all["Error Message"] == "")]
            
            app_logger.info(f"Audit: Verified: {len(df_verified)}, Manual Review: {len(df_manual)}, Failed: {len(df_failed)}")
            
            total_filtered = len(df_verified) + len(df_manual) + len(df_failed)
            if total_filtered != len(shops):
                raise RuntimeError(f"Excel Writer Error: Row mismatch! Input ({len(shops)}) != Verified+Manual+Failed ({total_filtered})")
            
            df_summary = pd.DataFrame([{
                "Total Processed": len(shops),
                "Verified / Likely": len(df_verified),
                "Manual Review": len(df_manual),
                "Failed (Errors)": len(df_failed)
            }])

            sheets_data = {"Verified": df_verified, "Manual Review": df_manual, "Failed": df_failed, "Summary": df_summary}
            
            for sheet_name, df in sheets_data.items():
                ws = wb.create_sheet(title=sheet_name)
                for row in dataframe_to_rows(df, index=False, header=True): ws.append(row)
                self._format_sheet(ws)
            
            output_file = os.path.join(self.output_path, "Updated_Jewellery_Shops.xlsx")
            wb.save(output_file)
            wb.close()
            app_logger.info(f"Successfully wrote output to {output_file}")
            
        except Exception as e:
            app_logger.error(f"Failed to write output Excel: {e}")
            raise
