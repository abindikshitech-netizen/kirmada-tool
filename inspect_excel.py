from openpyxl import load_workbook
import os
p = r'D:\kirmada-tool\output\Updated_Jewellery_Shops.xlsx'
print('exists', os.path.exists(p))
wb = load_workbook(p)
print('sheets', wb.sheetnames)
for ws in wb.worksheets:
    print(ws.title, ws.max_row, ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True):
        print(row)
